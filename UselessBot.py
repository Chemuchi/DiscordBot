import asyncio
import random
import discord
import openpyxl
import pytz
import yt_dlp as youtube_dl

from discord.ext import commands
from datetime import datetime

from discord.ext.commands import MissingRequiredArgument, CommandInvokeError

from tokenp import *
from Hangang import *
from Currency import *
from Translate import *
from Imgur import *

intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix="&",intents=intents)

embed_color = 0x7F7F7F
data_loading = '데이터를 가져오는중입니다.. 잠시만 기다려주세요...!'

@bot.event
async def on_ready():
    print('디스코드 로그인중..')
    print(f'{bot.user}로 로그인 되었습니다.')
    print(f'ID : {bot.user.id}')
    await bot.change_presence(status=discord.Status.online, activity=discord.Game('작동'))


@bot.command(aliases=['테스트'])
async def hello(ctx):
    print('Console : 정상작동중..')
    await ctx.reply(f'{ctx.author.mention} 님 안녕하세요!\n{bot.user}정상 작동중입니다.')

@bot.command(aliases=['강제종료'])
async def forceoff(ctx):
    await ctx.send('봇을 강제로 종료합니다.')
    await bot.close()

@bot.command(aliases=['봇정보','정보','명령어'])
async def commands(ctx):
    embed = discord.Embed(title="UselessBot 입니다.", description="개인서버 프로젝트용", color=embed_color)
    embed.set_thumbnail(url="https://pngimg.com/uploads/trash_can/trash_can_PNG18441.png")
    embed.add_field(name="🛠️서버관리", value="삭제", inline=False)
    embed.add_field(name="💰경제", value="등록, 출석, 내정보", inline=False)
    embed.add_field(name=":slot_machine:재미", value="가위바위보, 랜덤박스, 랜덤, imgur", inline=False)
    embed.add_field(name="🎸기타", value="번역, 환율, 환율계산", inline=False)
    embed.add_field(name="", value=" ", inline=False)
    embed.add_field(name="💻Github", value='[https://github.com/Chemuchi/DiscordBot]', inline=False)
    await ctx.send(embed=embed)

'''@bot.command(aliases=['사용법','help'])
async def command_help(ctx,arg):
    if arg=='삭제'
    '''
'''----------------------------------------------유저관련---------------------------------------------------'''
@bot.command(aliases=['등록'])
async def register(ctx):
    tz = pytz.timezone('Asia/Seoul')
    now = datetime.now(tz)
    user = ctx.author
    name = user.name
    user_id = hex(user.id)
    money = int(5000)
    last_checkin = hex(int(now.timestamp()))

    # Excel 파일 불러오거나 없으면 UserDB 생성
    try:
        wb = openpyxl.load_workbook('userDB.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['이름','ID','돈','마지막 출석 시간'])

    # 중복값 확인
    for row in sheet.iter_rows(values_only=True):
        if row[1] == user_id:
            embed = discord.Embed(title="✅유저 등록", description=" \n", color=embed_color)
            embed.add_field(name='이미 등록되어있는 유저입니다 ', value=" ", inline=False)
            embed.set_footer(text=ctx.author.name, icon_url=ctx.author.display_avatar)
            await ctx.reply(embed=embed)

            return
    # 사용자 정보를 Excel 파일에 추가
    sheet.append([name, hex(user.id), money,last_checkin])
    wb.save('userDB.xlsx')
    embed = discord.Embed(title="✅유저 등록", description="", color=embed_color)
    embed.add_field(name=ctx.author.name, value=ctx.author.id, inline=False)
    embed.set_thumbnail(url=ctx.author.display_avatar)
    embed.set_footer(text=ctx.author.name, icon_url=ctx.author.display_avatar)
    await ctx.reply(embed=embed)

@bot.command(aliases=['내정보'])
async def user_info(ctx):
    user = ctx.author
    date = datetime.utcfromtimestamp(((int(user.id) >> 22) + 1420070400000) / 1000)
    year = str(date.year)
    month = str(date.month)
    day = str(date.day)
    wb = openpyxl.load_workbook('userDB.xlsx')
    sheet = wb.active
    money = 0
    for row in sheet.iter_rows(values_only=True):
        if row[1] == hex(user.id):
            money = row[2]
            break
    embed = discord.Embed(title="🪪유저정보", color=embed_color)
    embed.add_field(name=f'{user.display_name}', value="", inline=False)
    embed.add_field(name="디스코드 가입일", value=year + '년 ' + month + '월 ' + day + "일 ", inline=True)
    embed.add_field(name="소지금", value=str(money) + "원", inline=True)
    embed.set_image(url=user.display_avatar)
    await ctx.reply(embed=embed)

@bot.command(aliases=['출석'])
async def checkin(ctx):
    tz = pytz.timezone('Asia/Seoul')
    now = datetime.now(tz)
    user = ctx.author
    wb = openpyxl.load_workbook('userDB.xlsx')
    sheet = wb.active
    last_checkin = None

    for row in sheet.iter_rows(values_only=True):
        if row[1] == hex(user.id):
            last_checkin = row[3]
            break

    if last_checkin is None:
        for row in sheet.iter_rows():
            if row[1].value == hex(user.id):
                row[2].value += 5000
                row[3].value = hex(int(now.timestamp()))
                wb.save('userDB.xlsx')
                break
        embed = discord.Embed(title=":gift:일일 출석", description="", color=embed_color)
        embed.add_field(name=f"{user.name}님 출석 완료", value="", inline=False)
        embed.set_thumbnail(url=ctx.author.display_avatar)
        embed.set_footer(text=ctx.author.name, icon_url=ctx.author.display_avatar)
        await ctx.reply(embed=embed)
        return
    last_checkin_date = datetime.fromtimestamp(int(last_checkin, 16)).date()
    current_date = now.date()
    if current_date > last_checkin_date:
        for row in sheet.iter_rows():
            if row[1].value == hex(user.id):
                row[2].value += 5000
                row[3].value = hex(int(now.timestamp()))
                wb.save('userDB.xlsx')
                break
        embed = discord.Embed(title=":gift:일일 출석", description="", color=embed_color)
        embed.add_field(name=f"{user.name}님 출석 완료", value="", inline=False)
        embed.set_thumbnail(url=ctx.author.display_avatar)
        embed.set_footer(text=ctx.author.name, icon_url=ctx.author.display_avatar)
        await ctx.reply(embed=embed)
    else:
        embed = discord.Embed(title=":gift:일일 출석", description="", color=embed_color)
        embed.add_field(name=f"{user.name}님은 이미 출석 하셨습니다.", value='', inline=False)
        embed.set_thumbnail(url=ctx.author.display_avatar)
        embed.set_footer(text=ctx.author.name, icon_url=ctx.author.display_avatar)
        await ctx.reply(embed=embed)


'''-------------------------------------------------------------------------------------------------'''

'''-----------------------------------------------서버관리--------------------------------------------------'''

@bot.command(aliases=['삭제'])
async def delete(ctx,amount : int):
    embed = discord.Embed(title="✂️메세지 삭제", description="", color=embed_color)
    role = discord.utils.get(ctx.guild.roles, name='서버관리')
    if role in ctx.author.roles:
        embed.add_field(name=f'정말로 삭제하시겠습니까? 이 행동은 되돌릴 수 없습니다!!', value=f'{amount}만큼의 메세지가 삭제됩니다.', inline=False)
        bot_message = await ctx.reply(embed=embed)

        await bot_message.add_reaction('✅')

        def check(reaction, user):
            return user == ctx.author and str(reaction.emoji) == '✅'

        try:
            reaction, user = await bot.wait_for('reaction_add', timeout=5.0, check=check)
        except asyncio.TimeoutError:
            embed = discord.Embed(title="✂️메세지 삭제", description=" ", color=embed_color)
            embed.add_field(name="삭제가 취소되었습니다.", value=" ")
            await bot_message.clear_reactions()
            await bot_message.edit(embed=embed)
        else:
            deleted_messages = await ctx.channel.purge(limit=amount+2)
            embed = discord.Embed(title="✂️메세지 삭제", description=" ", color=embed_color)
            embed.add_field(name=f"{amount}개의 메세지를 삭제 완료했습니다.", value=f"삭제된 메세지 로그는 📜 이모지를 눌러주세요.", inline=False)
            embed.set_footer(text=f'삭제자 : {ctx.author}', icon_url=ctx.author.display_avatar)
            log_message = await ctx.send(embed=embed)
            await log_message.add_reaction('📜')

            def check(reaction, user):
                return user == ctx.author and str(reaction.emoji) == '📜'
            try:
                reaction, user = await bot.wait_for('reaction_add', timeout=5.0, check=check)
            except asyncio.TimeoutError:
                embed.clear_fields()
                embed.add_field(name=f"{amount}개의 메세지를 삭제 완료했습니다.", value=f"", inline=False)
                await log_message.edit(embed=embed)
                await log_message.clear_reactions()
            else:
                log_message = '\n'.join([f'{ctx.author}: {message.content}' for message in reversed(deleted_messages)])
                await ctx.author.send(f'요청하신 메세지 로그입니다.\n\n{log_message}')

    else:
        embed = discord.Embed(title="✂️메세지 삭제", description=" ", color=embed_color)
        embed.add_field(name="이 명령어를 실행할 권한이 없습니다.", value=f" ",inline=False)
        await ctx.reply(embed=embed)
@delete.error
async def delete_error(ctx, error):
    if isinstance(error, MissingRequiredArgument):
        embed = discord.Embed(title="✂️메세지 삭제", description=" ", color=embed_color)
        embed.add_field(name='삭제할 메세지 수를 입력해주세요.',value='명렁어+봇답장 총 n+2 의 메세지가 삭제됩니다.',inline=False)
        await ctx.reply(embed=embed)

'''-------------------------------------------------------------------------------------------------'''

'''---------------------------------------------재미기능----------------------------------------------------'''

@bot.command(aliases=['가위바위보'])
async def rock_paper_scissors(ctx, bet_money : int):
    user = ctx.author
    wb = openpyxl.load_workbook('userDB.xlsx')
    sheet = wb.active
    money = 0
    for row in sheet.iter_rows(values_only=True):
        if row[1] == hex(user.id):
            money = row[2]
            break
    if bet_money > int(50000):
        embed = discord.Embed(title="가위바위보", description="", color=embed_color)
        embed.add_field(name='50000원을 초과해서 베팅할수는 없습니다!', value=f'{user.name}님의 전재산 : {money}', inline=False)
        await ctx.reply(embed=embed)
        return
    if bet_money > money:
        embed = discord.Embed(title="가위바위보", description="", color=embed_color)
        embed.add_field(name='돈이 부족합니다!', value=f'{user.name}님의 전재산 : {money}', inline=False)
        await ctx.reply(embed=embed)
        return
    if bet_money == 0:
        embed = discord.Embed(title="가위바위보", description="", color=embed_color)
        embed.add_field(name='0원을 베팅할수는 없습니다!', value=f'{user.name}님의 전재산 : {money}', inline=False)
        await ctx.reply(embed=embed)
        return
    if bet_money == None:
        embed = discord.Embed(title="가위바위보", description="", color=embed_color)
        embed.add_field(name='1 이상의 돈을 걸어야합니다!\n사용법 : $가위바위보 (금액)', value=f'{user.name}님의 전재산 : {money}', inline=False)
        await ctx.reply(embed=embed)

    choices = ['✊', '✌️', '🖐️']
    bot_choice = random.choice(choices)
    embed = discord.Embed(title="가위바위보", description="10초 내로 선택하세요!", color=embed_color)
    embed.add_field(name="주먹", value="✊", inline=False)
    embed.add_field(name="가위", value="✌️", inline=False)
    embed.add_field(name="보", value="🖐️", inline=False)
    sent_message = await ctx.reply(embed=embed)
    for choice in choices:
        await sent_message.add_reaction(choice)

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in choices

    try:
        reaction, user = await bot.wait_for('reaction_add', timeout=10.0, check=check)
    except asyncio.TimeoutError:
        embed = discord.Embed(title="가위바위보", description=" ", color=embed_color)
        embed.add_field(name='시간초과!', value="", inline=False)
        await sent_message.edit(embed=embed)
        await sent_message.clear_reactions()
    else:
        if str(reaction.emoji) == bot_choice:
            for row in sheet.iter_rows(values_only=True):
                if row[1] == hex(user.id):
                    money = row[2]
                    wb.save("userDB.xlsx")
                    break
            embed = discord.Embed(title="가위바위보", description=" ", color=embed_color)
            embed.add_field(name=f"비겼습니다! 봇의 선택 : {bot_choice} {user.name}님의 선택 : {reaction.emoji}", value=f"{user.name}님의 남은 돈 : {money}원", inline=False)
            await sent_message.edit(embed=embed)
        elif (str(reaction.emoji) == '✊' and bot_choice == '✌️') or (
                str(reaction.emoji) == '✌️' and bot_choice == '🖐️') or (
                str(reaction.emoji) == '🖐️' and bot_choice == '✊'):
            embed = discord.Embed(title="가위바위보", description=" ", color=0x7F7F7F)
            for row in sheet.iter_rows():
                if row[1].value == hex(user.id):
                    row[2].value += bet_money
                    money += bet_money
                    wb.save("userDB.xlsx")
                    break
            embed.add_field(name=f"이겼습니다! 봇의 선택 : {bot_choice} {user.name}님의 선택 : {reaction.emoji}" , value=f"{user.name}님의 돈 : {money}원", inline=False)
            await sent_message.edit(embed=embed)
            await sent_message.clear_reactions()

        else:
            for row in sheet.iter_rows():
                if row[1].value == hex(user.id):
                    row[2].value -= bet_money
                    money -= bet_money
                    wb.save("userDB.xlsx")
                    break
            embed = discord.Embed(title="가위바위보", description=" ", color=0x7F7F7F)
            embed.add_field(name=f"졌습니다! 봇의 선택 : {bot_choice} {user.name}님의 선택 : {reaction.emoji}", value=f"{user.name}님의 돈 : {money}원", inline=False)
            await sent_message.edit(embed=embed)
            await sent_message.clear_reactions()
@rock_paper_scissors.error
async def rpc_error(ctx,error):
    user = ctx.author
    wb = openpyxl.load_workbook('userDB.xlsx')
    sheet = wb.active
    money = 0
    for row in sheet.iter_rows(values_only=True):
        if row[1] == hex(user.id):
            money = row[2]
            break
    if isinstance(error, MissingRequiredArgument):
        embed = discord.Embed(title="가위바위보", description="", color=embed_color)
        embed.add_field(name='돈을 걸어야합니다!', value=f'{user.name}님의 돈 : {money}', inline=False)
        await ctx.reply(embed=embed)

@bot.command(aliases=['랜덤박스'])
async def randombox(ctx):
    user = ctx.author

    wb = openpyxl.load_workbook('userDB.xlsx')
    sheet = wb.active

    money = None
    for row in sheet.iter_rows(values_only=True):
        if row[1] == hex(user.id):
            money = row[2]
            break
    if money < 1000:
        embed = discord.Embed(title=":moneybag: 랜덤박스", description="", color=embed_color)
        embed.add_field(name='돈이 부족합니다! (1000원)', value=f'{user.name}님의 전재산  : {money}', inline=False)
        await ctx.reply(embed=embed)
        return
    else:
        for row in sheet.iter_rows():
            if row[1].value == hex(user.id):
                row[2].value -= 1000
                wb.save("userDB.xlsx")
                break
        number_emojis = ['1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣']
        money_list = [500, 1000, 5000, 10000, 20000]
        money_probability = [0.4, 0.8, 0.2, 0.01, 0.008]

        embed = discord.Embed(title=":moneybag: 랜덤박스", description="", color=embed_color)
        embed.add_field(name='이모지 하나를 선택해주세요!', value='', inline=False)
        sent_message = await ctx.reply(embed=embed)

        for emoji in number_emojis:
            await sent_message.add_reaction(emoji)

        emoji_money_dict = {}

        for emoji in number_emojis:
            money = random.choices(money_list, weights = money_probability)[0]
            emoji_money_dict[emoji] = money

        def check(reaction, user):
            return user == ctx.author and str(reaction.emoji) in number_emojis

        try:
            reaction, user = await bot.wait_for('reaction_add', timeout=10.0, check=check)
        except asyncio.TimeoutError:
            embed = discord.Embed(title=":moneybag: 랜덤박스", description=" ", color=embed_color)
            embed.add_field(name='시간초과!', value="", inline=False)
            await sent_message.edit(embed=embed)
            await sent_message.clear_reactions()
        else:
            selected_emoji = str(reaction.emoji)
            selected_money = emoji_money_dict[selected_emoji]
            for row in sheet.iter_rows():
                if row[1].value == hex(user.id):
                    row[2].value += selected_money
                    wb.save("userDB.xlsx")
                    break
            money = None

            for row in sheet.iter_rows(values_only=True):
                if row[1] == hex(user.id):
                    money = row[2]
                    wb.save("userDB.xlsx")
                    break
            embed = discord.Embed(title=":moneybag: 랜덤박스", description=" ", color=embed_color)
            embed.add_field(name=f"{selected_money}원 당첨!", value=f"{user.name}의  전재산 : {money}원", inline=False)
            await sent_message.edit(embed=embed)
            await sent_message.clear_reactions()

@bot.command(aliases=['랜덤'])
async def imgur_random_word(ctx):
    image_url = get_random_image(random_words())
    search_word = str(random_words())
    await ctx.reply(image_url)

@bot.command(aliases=['imgur'])
async def imgur_random_image(ctx,*args):
    text = ' '.join(args)
    image_url = get_random_image(text)
    await ctx.reply(image_url)
@imgur_random_image.error
async def iri_error(ctx, error):
    if isinstance(error, CommandInvokeError):
        embed = discord.Embed(title="🔎 Imgur", description=" ", color=embed_color)
        embed.add_field(name='검색어를 입력해주세요.',value=' ',inline=False)
        await ctx.reply(embed=embed)




'''-----------------------------------------------------------------------------------------'''
'''-----------------------------------------정보-----------------------------------------------'''

@bot.command(aliases=['한강'])
async def hangang(ctx):
    embed = discord.Embed(title="🌡️한강 물 온도", description="", color=embed_color)
    embed.add_field(name=data_loading, value="", inline=False)
    sent_message = await ctx.reply(embed=embed)
    embed.clear_fields()
    embed.add_field(name=f"현재 한강의 온도는 ", value=f"{temp()}입니다.", inline=False)
    await sent_message.edit(embed=embed)

@bot.command(aliases=['환율'])
async def exchange(ctx):
    flags = ['🇺🇸', '🇯🇵', '🇬🇧', '🇪🇺', '🇹🇷']
    embed = discord.Embed(title=":currency_exchange: 환율", description="", color=embed_color)
    embed.add_field(name='나라를 선택해주세요!', value="", inline=False)
    sent_message = await ctx.reply(embed=embed)
    for flag in flags:
        await sent_message.add_reaction(flag)

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in flags

    try:
        reaction, user = await bot.wait_for('reaction_add', timeout=7.0, check=check)
    except asyncio.TimeoutError:
        embed.clear_fields()
        embed.add_field(name='시간이 초과되었습니다.',value='',inline=False)
        await sent_message.edit(embed=embed)
        await sent_message.clear_reactions()
    else:
        embed.clear_fields()
        embed.add_field(name=data_loading, value='',inline=False)
        await sent_message.edit(embed=embed)
        if str(reaction.emoji) == '🇺🇸':
            embed.clear_fields()
            embed.add_field(name=f'1 달러는 {US()}원 입니다.', value='',inline=False)
            await sent_message.edit(embed=embed)
            pass
        elif str(reaction.emoji) == '🇯🇵':
            embed.clear_fields()
            embed.add_field(name=f'100 엔은 {JP()}원 입니다.', value='', inline=False)
            await sent_message.edit(embed=embed)
            pass
        elif str(reaction.emoji) == '🇬🇧':
            embed.clear_fields()
            embed.add_field(name=f'1 파운드는 {GB()}원 입니다.', value='', inline=False)
            await sent_message.edit(embed=embed)
            pass
        elif str(reaction.emoji) == '🇪🇺':
            embed.clear_fields()
            embed.add_field(name=f'1 유로는 {EU()}원 입니다.', value='', inline=False)
            await sent_message.edit(embed=embed)
            pass
        elif str(reaction.emoji) == '🇹🇷':
            embed.clear_fields()
            embed.add_field(name=f'1 리라는 {TR()}원 입니다.', value='', inline=False)
            await sent_message.edit(embed=embed)
            pass
        await sent_message.clear_reactions()
@bot.command(aliases=['환율계산'])
async def exchange_calc(ctx, amount : float):
    flags = ['🇺🇸', '🇯🇵', '🇬🇧', '🇪🇺', '🇹🇷']
    choices = ['1️⃣', '2️⃣']
    embed = discord.Embed(title=":currency_exchange: 환율 계산", description="", color=embed_color)
    embed.add_field(name='계산 옵션을 선택해주세요.', value="원 에서 외화는 1️⃣, 외화 에서 원은 2️⃣ 를 눌러주세요.", inline=False)
    sent_message = await ctx.reply(embed=embed)
    for choice in choices:
        await sent_message.add_reaction(choice)

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in choices

    try:
        reaction, user = await bot.wait_for('reaction_add', timeout=10.0, check=check)
    except asyncio.TimeoutError:
        embed.clear_fields()
        embed.add_field(name='시간이 초과되었습니다.',value='', inline=False)
        await sent_message.edit(embed=embed)
        await sent_message.clear_reactions()
    else:
        if str(reaction.emoji) == '1️⃣':
            await sent_message.clear_reactions()
            embed.clear_fields()
            embed.add_field(name='나라를 선택해주세요!', value="", inline=False)
            await sent_message.edit(embed=embed)
            for flag in flags:
                await sent_message.add_reaction(flag)

            def check(reaction, user):
                return user == ctx.author and str(reaction.emoji) in flags

            try:
                reaction, user = await bot.wait_for('reaction_add', timeout=7.0, check=check)
            except asyncio.TimeoutError:
                embed.clear_fields()
                embed.add_field(name='시간이 초과되었습니다.', value='', inline=False)
                await sent_message.edit(embed=embed)
                await sent_message.clear_reactions()
            else:
                embed.clear_fields()
                embed.add_field(name=data_loading, value='', inline=False)
                await sent_message.edit(embed=embed)
                if str(reaction.emoji) == '🇺🇸':
                    embed.clear_fields()
                    formatted_value = f'{amount / US():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}원은 약 {formatted_value}달러 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇯🇵':
                    embed.clear_fields()
                    formatted_value = f'{(amount * 100) / JP():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}원은 약 {formatted_value}엔 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇬🇧':
                    embed.clear_fields()
                    formatted_value = f'{amount / GB():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}원은 약 {formatted_value}파운드 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇪🇺':
                    embed.clear_fields()
                    formatted_value = f'{amount / EU():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}원은 약 {formatted_value}유로 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇹🇷':
                    embed.clear_fields()
                    formatted_value = f'{amount / TR():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}원은 약 {formatted_value}리라 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                await sent_message.clear_reactions()
        elif str(reaction.emoji) == '2️⃣':
            await sent_message.clear_reactions()
            embed.clear_fields()
            embed.add_field(name='나라를 선택해주세요!', value="", inline=False)
            await sent_message.edit(embed=embed)
            for flag in flags:
                await sent_message.add_reaction(flag)

            def check(reaction, user):
                return user == ctx.author and str(reaction.emoji) in flags

            try:
                reaction, user = await bot.wait_for('reaction_add', timeout=7.0, check=check)
            except asyncio.TimeoutError:
                embed.clear_fields()
                embed.add_field(name='시간이 초과되었습니다.', value='', inline=False)
                await sent_message.edit(embed=embed)
                await sent_message.clear_reactions()
            else:
                embed.clear_fields()
                embed.add_field(name=data_loading, value='', inline=False)
                await sent_message.edit(embed=embed)
                if str(reaction.emoji) == '🇺🇸':
                    embed.clear_fields()
                    formatted_value = f'{amount * US():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}달러는 약 {formatted_value}원 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇯🇵':
                    embed.clear_fields()
                    formatted_value = f'{(amount * JP()) / 100:,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}엔은 약 {formatted_value}원 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇬🇧':
                    embed.clear_fields()
                    formatted_value = f'{amount * GB():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}파운드는 약 {formatted_value}원 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇪🇺':
                    embed.clear_fields()
                    formatted_value = f'{amount * EU():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}유로는 약 {formatted_value}원 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                elif str(reaction.emoji) == '🇹🇷':
                    embed.clear_fields()
                    formatted_value = f'{amount * TR():,.2f}'.rstrip('0').rstrip('.')
                    embed.add_field(name=f'{amount}리라는 약 {formatted_value}원 입니다.', value='', inline=False)
                    await sent_message.edit(embed=embed)
                await sent_message.clear_reactions()
@exchange_calc.error
async def ec_error(ctx, error):
    if isinstance(error, MissingRequiredArgument):
        embed = discord.Embed(title=":currency_exchange: 환율 계산", description=" ", color=embed_color)
        embed.add_field(name='계산할 금액을 입력해주세요.', value=' ', inline=False)
        await ctx.reply(embed=embed)

@bot.command(aliases=['번역'])
async def translator(ctx,*args):
    embed = discord.Embed(title=":pager: 번역", description="", color=embed_color)
    embed.set_footer(text='번역 제공 : Naver Papago')
    if not args:
        embed.add_field(name='번역할 단어 또는 문장을 입력해주세요.',value='',inline=False)
        await ctx.reply(embed=embed)
        return
    text = ' '.join(args)
    flags = ['🇺🇸', '🇯🇵', '🇰🇷', '🇨🇳', '🇷🇺']
    embed.set_footer(text='번역 제공 : Naver Papago')
    embed.add_field(name='나라를 선택해주세요!', value="선택하신 나라의 언어로 번역됩니다.", inline=False)
    sent_message = await ctx.reply(embed=embed)
    for flag in flags:
        await sent_message.add_reaction(flag)

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in flags

    try:
        reaction, user = await bot.wait_for('reaction_add', timeout=7.0, check=check)
    except asyncio.TimeoutError:
        embed.clear_fields()
        embed.add_field(name='시간이 초과되었습니다.', value='', inline=False)
        await sent_message.edit(embed=embed)
        await sent_message.clear_reactions()
    else:
        embed.clear_fields()
        await sent_message.edit(embed=embed)

        if str(reaction.emoji) == '🇺🇸':
            embed.add_field(name=f'{translate(text,"en")}', value='', inline=False)
            pass
        elif str(reaction.emoji) == '🇯🇵':
            embed.add_field(name=f'{translate(text, "ja")}', value='', inline=False)
            pass
        elif str(reaction.emoji) == '🇰🇷':
            embed.add_field(name=f'{translate(text, "ko")}', value='', inline=False)
            pass
        elif str(reaction.emoji) == '🇨🇳':
            embed.add_field(name=f'{translate(text, "zh-CN")}', value='', inline=False)
            pass
        elif str(reaction.emoji) == '🇷🇺':
            embed.add_field(name=f'{translate(text, "ru")}', value='', inline=False)
            pass

        await sent_message.edit(embed=embed)
        await sent_message.clear_reactions()

'''-------------------------------------------노래---------------------------------------------------'''






'''-------------------------------------------------------------------------------------------------'''

bot.run(token1())
