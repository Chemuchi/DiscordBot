from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4

default_money = 10000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def checkRow():
    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

def signup(_name, _id):
    _row = checkRow()

    ws.cell(row=2, column=c_name, value=_name)
    ws.cell(row=2, column=c_id, value =_id)
    ws.cell(row=2, column=c_money, value = default_money)
    ws.cell(row=2, column=c_lvl, value = 1)

    wb.save("userDB.xlsx")
def checkName(_name, _id):
    for row in range(2, ws.max_row+1):
        if ws.cell(row,1).value == _name and ws.cell(row,2).value == _id:
            break
            return False
        else:
            return True
            break
def delete():
    ws.delete_rows(2,ws.max_row)
    wb.save("userDB.xlsx")

'''def userInfo(_name, _id):
    if not checkName(_name, _id):
        for row in range(2, ws.max_row+2):
    	    if ws.cell(row, 1).value == _name and ws.cell(row, 2).value == _id:
                return ws.cell(row,1).value, ws.cell(row,2).value
                break
    else:
    	return None, None'''