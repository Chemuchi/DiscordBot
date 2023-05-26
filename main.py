import discord
import random
import datetime
import openpyxl
import asyncio
import pytz
from tokenp import *
from datetime import datetime



token = token1()

intents = discord.Intents.default()
intents.members = True
intents.message_content = True


client = discord.Client(intents=intents)
bot = discord.Client(intents=intents)




@client.event
async def on_ready():
    print(f'{client.user}실행 완료')


@client.event
async def on_message(message):
    if message.author == client.user:
        return


@client.event

async def on_message(message):
    if message.author == client.user:
        return

    if message.content.startswith("$가위바위보"):
        user = message.author

        wb = openpyxl.load_workbook("userDB.xlsx")
        sheet = wb.active

        money = None
        for row in sheet.iter_rows(values_only=True):
            if row[1] == hex(user.id):
                money = row[2]
                break
        try:
            bet_money = int(message.content.split()[1])
        except IndexError:
            embed = discord.Embed(title="가위바위보", description="", color=0xC19D25)
            embed.add_field(name='돈을 걸어야합니다!\n사용법 : $가위바위보 (금액)', value=f'{user.name}님의 돈 : {money}', inline=False)
            await message.channel.send(embed=embed,reference = message)
            return

        if bet_money > money:
            embed = discord.Embed(title="가위바위보", description="", color=0xC19D25)
            embed.add_field(name='돈이 부족합니다!',value=f'{user.name}님의 전재산 : {money}',inline=False)
            await message.channel.send(embed=embed, reference=message)
            return
        if bet_money == 0:
            embed = discord.Embed(title="가위바위보", description="", color=0xC19D25)
            embed.add_field(name='0원을 베팅할수는 없습니다!', value=f'{user.name}님의 전재산 : {money}', inline=False)
            await message.channel.send(embed=embed, reference=message)
            return
        if bet_money > 50001:
            embed = discord.Embed(title="가위바위보", description="", color=0xC19D25)
            embed.add_field(name='50000원을 초과해서 베팅할수는 없습니다!', value=f'{user.name}님의 전재산 : {money}', inline=False)
            await message.channel.send(embed=embed, reference=message)
            return


        choices = ['✊', '✌️', '🖐️']
        bot_choice = random.choice(choices)
        embed = discord.Embed(title="가위바위보",description="10초 내로 선택하세요!",color = 0xC19D25)
        embed.add_field(name="주먹", value="✊", inline=True)
        embed.add_field(name="가위", value="✌️", inline=True)
        embed.add_field(name="보", value="🖐️", inline=True)
        sent_message = await message.channel.send(embed=embed,reference = message)
        for choice in choices:
            await sent_message.add_reaction(choice)
        def check(reaction, user):
            return user == message.author and str(reaction.emoji) in choices

        try:
            reaction, user = await client.wait_for('reaction_add',timeout = 10.0, check=check)
        except asyncio.TimeoutError:
            embed = discord.Embed(title="가위바위보",description=" ",color = 0xC19D25)
            embed.add_field(name='시간초과!',value="",inline=False)
            await message.channel.send(embed=embed,reference = message)
        else:
            if str(reaction.emoji) == bot_choice:
                for row in sheet.iter_rows(values_only=True):
                    if row[1] == hex(user.id):
                        money = row[2]
                        wb.save("userDB.xlsx")
                        break
                embed = discord.Embed(title="가위바위보", description=" ", color=0xC19D25)
                embed.add_field(name = f"비겼습니다! 봇의 선택: {bot_choice}",value=f"{user.name}님의 남은 돈 : {money}원",inline=False)
                await message.channel.send(embed=embed,reference = message)
            elif (str(reaction.emoji) == '✊' and bot_choice == '✌️') or (
                    str(reaction.emoji) == '✌️' and bot_choice == '🖐️') or (
                    str(reaction.emoji) == '🖐️' and bot_choice == '✊'):
                embed = discord.Embed(title="가위바위보", description=" ", color=0xC19D25)
                for row in sheet.iter_rows():
                    if row[1].value == hex(user.id):
                        row[2].value += bet_money
                        money += bet_money
                        wb.save("userDB.xlsx")
                        break
                embed.add_field(name=f"이겼습니다! 봇의 선택: {bot_choice}",value=f"{user.name}님의 돈 : {money}원",inline=False)
                await message.channel.send(embed=embed, reference=message)

            else:
                for row in sheet.iter_rows():
                    if row[1].value == hex(user.id):
                        row[2].value -= bet_money
                        money -= bet_money
                        wb.save("userDB.xlsx")
                        break
                embed = discord.Embed(title="가위바위보", description=" ", color=0xC19D25)
                embed.add_field(name=f"졌습니다! 봇의 선택: {bot_choice}",value=f"{user.name}님의 돈 : {money}원",inline=False)
                await message.channel.send(embed=embed, reference=message)



    #내정보 (이름,가입일자,아바타)
    if message.content == '$내정보':
        user = message.author
        date = datetime.utcfromtimestamp(((int(user.id) >> 22) + 1420070400000) / 1000)
        year = str(date.year)
        month = str(date.month)
        day = str(date.day)
        wb = openpyxl.load_workbook('userDB.xlsx')
        sheet = wb.active
        money = 1
        for row in sheet.iter_rows(values_only=True):
            if row[1] == hex(user.id):
                money = row[2]
                break
        embed = discord.Embed(title="유저정보",color =0x9CFF58)
        embed.add_field(name = user,value = "", inline = False)
        embed.add_field(name="디스코드 가입일", value=year+'년 '+month+'월 '+day+"일 ", inline=True)
        embed.add_field(name="소지금", value=str(money) + "원", inline=True)
        embed.set_image(url=user.display_avatar)
        await message.channel.send(embed=embed,reference=message)

    if message.content == "$등록":
        tz = pytz.timezone('Asia/Seoul')
        now = datetime.now(tz)
        user = message.author
        name = user.name  # 사용자 정보 가져오기
        user_id = hex(user.id)
        money = int(30000)  # 기본값
        last_checkin = hex(int(now.timestamp()))  # 날짜/시간 객체를 정수로 변환

        # Excel 파일 불러오기 또는 생성
        try:
            wb = openpyxl.load_workbook('userDB.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(['이름', 'ID', '돈', '마지막 출석 시간'])

        # 중복값 확인
        for row in sheet.iter_rows(values_only=True):
            if row[1] == user_id:
                embed = discord.Embed(title="유저 등록", description=" \n", color=0x3AE9E9)
                embed.add_field(name='이미 등록되어있는 유저입니다 ✅ ', value=" ", inline=False)
                embed.set_footer(text=message.author.name, icon_url=message.author.display_avatar)
                await message.channel.send(embed=embed, reference=message)

                return

        # 사용자 정보를 Excel 파일에 추가
        sheet.append([name, hex(user.id), money, last_checkin])
        wb.save('userDB.xlsx')
        embed = discord.Embed(title="유저 등록", description="", color=0x3AE9E9)
        embed.add_field(name=message.author.name, value=message.author.id, inline=False)
        embed.set_thumbnail(url=message.author.display_avatar)
        embed.set_footer(text=message.author.name, icon_url=message.author.display_avatar)
        await message.channel.send(embed=embed, reference=message)

    if message.content.startswith("$삭제"):
        role = discord.utils.get(message.guild.roles, name = "봇관리")
        if role in message.author.roles:
        # 숫자 파싱
            try:
                num_messages = int(message.content.split()[1])
            except (IndexError, ValueError):
                await message.channel.send("잘못된 입력입니다. $삭제 (number) 형식으로 입력해주세요.",reference = message)
                return

            # 봇이 메시지를 보냄
            bot_message = await message.channel.send("정말 삭제하시겠습니까? 기록으로 남게될것입니다.\n삭제하실려면 아래 이모지를 10초 이내로 클릭해주세요.",reference = message)

            # 체크 표시 이모지 추가
            await bot_message.add_reaction("✅")

            # 체크 표시 이모지 대기
            def check(reaction, user):
                return user == message.author and str(reaction.emoji) == "✅"

            try:
                reaction, user = await client.wait_for("reaction_add", timeout=5, check=check)
            except asyncio.TimeoutError:
                await message.channel.send("확인 시간이 초과 되었습니다.",reference = message)
            else:
                # 메시지 삭제
                deleted_messages = await message.channel.purge(limit = num_messages + 2)#명령어 + 봇메세지
                embed = discord.Embed(title="🚧대화내역 삭제🚧", description="", color=0xDF2E2E)
                embed.add_field(name=f"메세지 {num_messages}개가 삭제되었습니다!! ", value="10초 내로 📜 이모지 반응시 삭제된 메시지 로그를 DM 으로 확인할 수 있습니다.", inline=False)
                embed.set_footer(text=f"삭제자 : {message.author.name}", icon_url=message.author.display_avatar)
                log_message = await message.channel.send(embed = embed)
                await log_message.add_reaction('📜')
                def check(reaction, user):
                    return user == message.author and str(reaction.emoji) == '📜'
                try:
                    reaction, user = await client.wait_for("reaction_add", timeout = 10, check = check)
                except asyncio.TimeoutError:
                    pass
                else:
                    #DM으로 로그전송
                    log_text = '\n'.join([f"{message.author}: {message.content}\n" for message in reversed(deleted_messages)])
                    await message.author.send(f"요청하신 삭제된 메세지 로그입니다.\n\n{log_text}")

        else:
            await message.channel.send("이 명령어를 실행할 권한이 없습니다.",reference = message)


    if message.content == "$출석":
        tz = pytz.timezone('Asia/Seoul')
        now = datetime.now(tz)
        user = message.author
        wb = openpyxl.load_workbook("userDB.xlsx")
        sheet = wb.active

        # 사용자의 마지막 출석 시간을 가져옴
        last_checkin = None
        for row in sheet.iter_rows(values_only=True):
            if row[1] == hex(user.id):
                last_checkin = row[3]
                break

        # 사용자가 처음 출석하는 경우
        if last_checkin is None:
            for row in sheet.iter_rows():
                if row[1].value == hex(user.id):
                    row[2].value += 30000  # 돈 추가
                    row[3].value = hex(int(now.timestamp()))  # 출석 시간 업데이트
                    wb.save("userDB.xlsx")
                    break
            embed = discord.Embed(title=":gift:일일 출석", description="", color=0xDBDBDB)
            embed.add_field(name=f"{user.name}님 출석 완료", value="", inline=False)
            embed.set_thumbnail(url=message.author.display_avatar)
            embed.set_footer(text=message.author.name, icon_url=message.author.display_avatar)
            await message.channel.send(embed=embed, reference=message)
            return

        # 사용자가 마지막으로 출석한 날짜와 현재 날짜를 비교합니다
        last_checkin_date = datetime.fromtimestamp(int(last_checkin, 16)).date()
        current_date = now.date()
        if current_date > last_checkin_date:
            for row in sheet.iter_rows():
                if row[1].value == hex(user.id):
                    row[2].value += 30000  # 돈 추가
                    row[3].value = hex(int(now.timestamp()))  # 출석 시간 업데이트
                    wb.save("userDB.xlsx")
                    break
            embed = discord.Embed(title=":gift:일일 출석", description="", color=0xDBDBDB)
            embed.add_field(name=f"{user.name}님 출석 완료", value="", inline=False)
            embed.set_thumbnail(url=message.author.display_avatar)
            embed.set_footer(text=message.author.name, icon_url=message.author.display_avatar)
            await message.channel.send(embed=embed, reference=message)
        else:
            embed = discord.Embed(title=":gift:일일 출석", description="", color=0xDBDBDB)
            embed.add_field(name=f"{user.name}님은 이미 출석 하셨습니다.", value="", inline=False)
            embed.set_thumbnail(url=message.author.display_avatar)
            embed.set_footer(text=message.author.name, icon_url=message.author.display_avatar)
            await message.channel.send(embed=embed, reference=message)

    if message.content == "$랜덤박스":
        user = message.author

        wb = openpyxl.load_workbook("userDB.xlsx")
        sheet = wb.active

        money = None
        for row in sheet.iter_rows(values_only=True):
            if row[1] == hex(user.id):
                money = row[2]
                break
        if money < 10000:
            embed = discord.Embed(title="랜덤박스", description="", color=0xC64DCA)
            embed.add_field(name='돈이 부족합니다!',value=f'{user.name}님의 전재산 : {money}',inline=False)
            await message.channel.send(embed=embed, reference=message)
            return
    result = random.choices(price =[0,1000,5000,20000,50000,100000,300000,500000],per = [30, 30, 20, 10, 7, 2, 0.9, 0.1], k=1)[0]
    if result == 0:
        embed = discord.Embed(title='랜덤박스', description = '', color=0xC64DCA)
        embed.add_field(name=f'저런 운도없어라~ 꽝이네요~~')
        await message.channel.send(embed = embed,reference = message)
    else:
        embed = discord.Embed(title = '랜덤박스', description = '',color = 0xC64DCA)
        embed.add_field(name = f'축하합니다!',value = f'{result}원에 당첨되셨습니다!')
        await message.channel.send(embed = embed,reference = message)

    if message.content == "$랜덤박스":
        user = message.author

        wb = openpyxl.load_workbook("userDB.xlsx")
        sheet = wb.active

        money = None
        for row in sheet.iter_rows(values_only=True):
            if row[1] == hex(user.id):
                money = row[2]
                break
        if money < 10000:
            embed = discord.Embed(title="랜덤박스", description="", color=0xC64DCA)
            embed.add_field(name='돈이 부족합니다!',value=f'{user.name}님의 전재산 : {money}',inline=False)
            await message.channel.send(embed=embed, reference=message)
            return
        else:
            for row in sheet.iter_rows():
                if row[1].value == hex(user.id):
                    row[2].value -= 10000
                    wb.save("userDB.xlsx")
                    break
            # 숫자 이모지 리스트
            number_emojis = ['1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣']
            # 금액 확률 리스트
            money_probabilities = [0.4, 0.4, 0.3,0.2, 0.09, 0.07, 0.02, 0.009, 0.001]
            # 금액 리스트
            money_list = [0, 1000, 5000, 10000,25000, 35000, 50000, 80000, 100000]

            # 메시지 전송
            embed = discord.Embed(title="랜덤박스", description="", color=0xC64DCA)
            embed.add_field(name='이모지 하나를 선택해주세요!', value=f'', inline=False)

            sent_message = await message.channel.send(embed=embed, reference=message)


            # 이모지 추가
            for emoji in number_emojis:
                await sent_message.add_reaction(emoji)

            # 이모지에 할당된 금액 딕셔너리 초기화
            emoji_money_dict = {}

            # 각 이모지에 금액 할당
            for emoji in number_emojis:
                money = random.choices(money_list, weights=money_probabilities)[0]
                emoji_money_dict[emoji] = money

            def check(reaction, user):
                return user == message.author and str(reaction.emoji) in number_emojis

            try:
                reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
            except asyncio.TimeoutError:
                await message.channel.send('시간 초과')
            else:
                selected_emoji = str(reaction.emoji)
                selected_money = emoji_money_dict[selected_emoji]
                for row in sheet.iter_rows():
                    if row[1].value == hex(user.id):
                        row[2].value += int(selected_money)
                        wb.save("userDB.xlsx")
                        break
                money = None
                # 돈관련 함수
                for row in sheet.iter_rows(values_only=True):
                    if row[1] == hex(user.id):
                        money = row[2]
                        wb.save("userDB.xlsx")
                        break
                if selected_money == 0:
                    embed = discord.Embed(title="랜덤박스", description="", color=0xC64DCA)
                    embed.add_field(name='꽝입니다..', value=f'{user.name}님의 전재산 : {money}', inline=False)
                    await message.channel.send(embed=embed, reference=message)
                else:
                    embed = discord.Embed(title="랜덤박스", description="", color=0xC64DCA)
                    embed.add_field(name=f'{selected_money}에 당첨되었습니다!', value=f'{user.name}님의 전재산 : {money}', inline=False)
                    await message.channel.send(embed=embed, reference=message)



client.run(token)


